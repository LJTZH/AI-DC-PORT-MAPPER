"""Output writer: generates multi-sheet Excel workbook with i18n support."""

from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Connection
from .i18n import get_lang


# ── Styling ─────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, num_cols: int):
    """Apply header styling to the first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data(ws, num_rows: int, num_cols: int):
    """Apply styling to data rows."""
    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER


def _auto_width(ws, num_cols: int, min_width: int = 10, max_width: int = 30):
    """Auto-fit column widths."""
    for col in range(1, num_cols + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = max(min_width, min(max_len + 2, max_width))


# ── Sheet generators ────────────────────────────────────────────────────

def write_port_mapping_sheet(ws, connections: list[Connection], S: dict[str, str]):
    """Write Sheet 1: Port Mapping Table."""
    headers = [
        S["col_src_device"], S["col_src_rack"], S["col_src_ru"],
        S["col_src_port"], S["col_src_port_type"],
        S["col_dst_device"], S["col_dst_rack"], S["col_dst_ru"],
        S["col_dst_port"], S["col_dst_port_type"],
        S["col_cable_type"], S["col_cable_len_std"], S["col_cable_len_calc"],
        S["col_needs_transceiver"], S["col_transceiver_type"],
        S["col_notes"],
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for conn in connections:
        ws.append([
            conn.src_device, conn.src_rack, conn.src_ru,
            conn.src_port, conn.src_port_type,
            conn.dst_device, conn.dst_rack, conn.dst_ru,
            conn.dst_port, conn.dst_port_type,
            conn.cable_type, conn.cable_length_m, conn.calculated_length_m,
            S["yes"] if conn.needs_transceiver else S["no"],
            conn.transceiver_type if conn.transceiver_type else S["na"],
            conn.notes,
        ])

    _style_data(ws, len(connections), len(headers))
    _auto_width(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _bom_category(ct: str, S: dict[str, str]) -> str:
    if "DAC" in ct:
        return "DAC"
    if "AOC" in ct:
        return "AOC"
    if "Fiber" in ct:
        return S["cat_fiber"]
    if "Cat" in ct:
        return S["cat_copper"]
    return ""


def write_cable_bom_sheet(ws, connections: list[Connection], S: dict[str, str]):
    """Write Sheet 2: Cable BOM with per-length breakdown."""
    headers = [
        S["bom_col_cable_type"], S["bom_col_category"],
        S["bom_col_length"], S["bom_col_qty"], S["bom_col_subtotal_len"],
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    # Group by (cable_type, length)
    groups: dict[tuple[str, float], list[Connection]] = defaultdict(list)
    for conn in connections:
        groups[(conn.cable_type, conn.cable_length_m)].append(conn)

    sorted_keys = sorted(groups.keys(), key=lambda k: (k[0], k[1]))

    total_qty = 0
    total_length = 0.0
    current_type = None
    type_qty = 0
    type_length = 0.0
    type_category = ""

    for (cable_type, length) in sorted_keys:
        items = groups[(cable_type, length)]
        qty = len(items)
        subtotal = round(length * qty, 2)
        category = _bom_category(cable_type, S)
        total_qty += qty
        total_length += subtotal

        if current_type is not None and cable_type != current_type:
            ws.append([
                f"{current_type} {S['bom_subtotal']}", type_category, S["na"],
                type_qty, round(type_length, 2),
            ])
            row = ws.max_row
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).font = Font(italic=True)
            type_qty = 0
            type_length = 0.0

        current_type = cable_type
        type_category = category
        type_qty += qty
        type_length += subtotal

        ws.append([cable_type, category, length, qty, subtotal])

    # Last type summary
    if current_type is not None:
        ws.append([
            f"{current_type} {S['bom_subtotal']}", type_category, S["na"],
            type_qty, round(type_length, 2),
        ])
        row = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).font = Font(italic=True)

    # Grand total
    ws.append([S["bom_grand_total"], "", S["na"], total_qty, round(total_length, 2)])
    summary_row = ws.max_row
    for col in range(1, len(headers) + 1):
        ws.cell(row=summary_row, column=col).font = Font(bold=True)

    _style_data(ws, ws.max_row - 1, len(headers))
    _auto_width(ws, len(headers))
    ws.freeze_panes = "A2"


def write_transceiver_sheet(ws, connections: list[Connection], S: dict[str, str]):
    """Write Sheet 3: Optical Transceiver Summary."""
    headers = [
        S["tx_col_type"], S["tx_col_qty"],
        S["tx_col_connections"], S["tx_col_notes"],
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    groups: dict[str, list[Connection]] = defaultdict(list)
    for conn in connections:
        if conn.needs_transceiver and conn.transceiver_type:
            groups[conn.transceiver_type].append(conn)

    for tx_type, items in sorted(groups.items()):
        tx_count = len(items) * 2
        ws.append([
            tx_type,
            tx_count,
            S["tx_conn_fmt"].format(n=len(items)),
            S["tx_note_fmt"].format(type=tx_type),
        ])

    _style_data(ws, len(groups), len(headers))
    _auto_width(ws, len(headers))
    ws.freeze_panes = "A2"


def write_device_summary_sheet(ws, connections: list[Connection], S: dict[str, str]):
    """Write Sheet 4: Per-Device Connection Summary."""
    headers = [
        S["dev_col_name"], S["dev_col_total"],
        S["dev_col_uplink"], S["dev_col_downlink"],
        S["dev_col_port_types"],
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    device_conns: dict[str, dict] = defaultdict(lambda: {
        "total": 0, "uplink": 0, "downlink": 0, "port_types": set()
    })

    for conn in connections:
        for side, dev_name, direction in [
            ("src", conn.src_device, conn.src_port_direction),
            ("dst", conn.dst_device, conn.dst_port_direction),
        ]:
            stats = device_conns[dev_name]
            stats["total"] += 1
            if direction == "uplink":
                stats["uplink"] += 1
            elif direction == "downlink":
                stats["downlink"] += 1
            stats["port_types"].add(
                conn.src_port_type if side == "src" else conn.dst_port_type
            )

    for dev_name in sorted(device_conns.keys()):
        stats = device_conns[dev_name]
        ws.append([
            dev_name,
            stats["total"],
            stats["uplink"],
            stats["downlink"],
            ", ".join(sorted(stats["port_types"])),
        ])

    _style_data(ws, len(device_conns), len(headers))
    _auto_width(ws, len(headers))
    ws.freeze_panes = "A2"


# ── Main writer ─────────────────────────────────────────────────────────

def write_output(
    filepath: str,
    connections: list[Connection],
    lang: str = "zh",
) -> str:
    """Generate the multi-sheet output Excel file.

    Args:
        filepath: Output file path (.xlsx).
        connections: List of Connection objects.
        lang: Language code — "zh", "ja", "en" (default "zh").

    Returns:
        The output filepath.
    """
    S = get_lang(lang)
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet(S["sheet_port_mapping"])
    write_port_mapping_sheet(ws1, connections, S)

    ws2 = wb.create_sheet(S["sheet_cable_bom"])
    write_cable_bom_sheet(ws2, connections, S)

    ws3 = wb.create_sheet(S["sheet_transceiver"])
    write_transceiver_sheet(ws3, connections, S)

    ws4 = wb.create_sheet(S["sheet_device_summary"])
    write_device_summary_sheet(ws4, connections, S)

    wb.save(filepath)
    return filepath
